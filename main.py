import openpyxl
import dataclasses
import typing as t
import pathlib
from selenium import webdriver
from glob import glob
from PIL import Image
import threading

# CONST
WB = openpyxl.load_workbook("./Rozklad_24_25.xlsx")
SHEET_FOR_PARSE = WB["Розклад"]


@dataclasses.dataclass
class InfoAboutLesson:
    numerator: str | None
    numerator_audience: str | None
    denominator: str | None
    denominator_audience: str | None

    @classmethod
    def from_dict(cls, list_: list) -> t.Self:
        return cls(
            numerator=list_[0][0],
            numerator_audience=list_[0][1],
            denominator=list_[1][0],
            denominator_audience=list_[1][1],
        )


@dataclasses.dataclass
class InfoAboutDay:
    day: str
    first: InfoAboutLesson | None
    second: InfoAboutLesson | None
    third: InfoAboutLesson | None
    fourth: InfoAboutLesson | None
    fifth: InfoAboutLesson | None
    sixth: InfoAboutLesson | None

    @classmethod
    def from_dict(cls, list_: list[str], day: str) -> t.Self:
        return cls(
            day=day,
            first=InfoAboutLesson.from_dict(list_[0:2]),
            second=InfoAboutLesson.from_dict(list_[2:4]),
            third=InfoAboutLesson.from_dict(list_[4:6]),
            fourth=InfoAboutLesson.from_dict(list_[6:8]),
            fifth=InfoAboutLesson.from_dict(list_[8:10]),
            sixth=InfoAboutLesson.from_dict(list_[10:12]),
        )


@dataclasses.dataclass
class Schedule:
    group_name: str
    monday: InfoAboutDay | None
    tuesday: InfoAboutDay | None
    wednesday: InfoAboutDay | None
    thursday: InfoAboutDay | None
    friday: InfoAboutDay | None

    @classmethod
    def from_dict(cls, list_: list) -> t.Self:
        return cls(
            group_name=list_[0][0],
            monday=InfoAboutDay.from_dict(list_[1:13], "Понеділок"),
            tuesday=InfoAboutDay.from_dict(list_[13:25], "Вівторок"),
            wednesday=InfoAboutDay.from_dict(list_[25:37], "Середа"),
            thursday=InfoAboutDay.from_dict(list_[37:49], "Четвер"),
            friday=InfoAboutDay.from_dict(list_[49:61], "П'ятниця"),
        )


class Practice:
    def __init__(self, color):
        self.rowspan = 1
        self.colspan = 2
        self.color = color

    def __bool__(self):
        return True

    def add_rowspan(self):
        self.rowspan += 1

    def __str__(self):
        return f"<td class='end {self.color} practice_start' rowspan='{self.rowspan}' colspan='{self.colspan}'>Практика</td>"


def get_data_from_sheet(column: int) -> list[list[str]]:
    result = []
    column = column * 2 + 1
    start_row = 6
    end_row = 67

    for row in range(start_row, end_row):
        value1 = SHEET_FOR_PARSE.cell(row=row, column=column).value
        value2 = SHEET_FOR_PARSE.cell(row=row, column=column + 1).value
        result.append([value1, value2])
    return result


def transform_list_to_html_list(list_: list[list]) -> list[list[str]]:
    result = []
    practice = False
    for count, row in enumerate(list_):
        # 2 gray ... 2 white etc.
        color = "dark" if not ((count - 1) // 2) % 2 else "light"
        middle_clear = f"<td class='middle {color}'></td>"
        end_clear = f"<td class='end {color}'></td>"
        middle = f"<td class='middle {color}'>{row[0]}</td>"
        end = f"<td class='end {color}'>{row[1]}</td>"
        if count in [1, 13, 25, 37, 49, 61]:
            practice = False
        if count == 0:
            result.append([row[0].replace("-", ""), row[1]])
        elif row[0] == "Практика":
            practice = Practice(color)
            result.append([practice, ""])
        elif row[0] is not None and row[1] is not None:
            practice = False
            result.append([middle, end])
        elif row[0] is not None and row[1] is None:
            practice = False
            result.append([middle, end_clear])
        elif row[0] is None and row[1] is None:
            if practice:
                result.append(["", ""])
                practice.add_rowspan()
            else:
                result.append([middle_clear, end_clear])
        elif row[0] is None and row[1] is not None:
            practice = False
            result.append([middle_clear, end])
    return result


def cut_big_words(list_: list[list[str]]) -> list[list[str]]:
    change_on_it = {
        "Історія: Україна і світ": "Істо: Укр і світ",
        "Зарубіжна література": "Заруб. літерат.",
        "Мистецтво (Методика)": "Мистецтво (Мет.)",
        "Практикум реставрац робіт": "Практикум рест.",
        "Методика образотворчого мистецтва": "Метод. обра. мист.",
        "Людина і суспільтво": "Людина і сусп.",
        "Декоративно-прикладне мистецтво": "Декор.-прикладне мист.",
        "Пластична анатомія": "Пласт. анатомія",
        "Осн наук досліджень": "Осн наук досл.",
        "Діяльнісний підхід (лего)": "Діяльн. підхід (лего)",
        "Історія:Україна і світ": "Історія: Укр і світ",
        "Іноземна мова/Ритміка": "Іноз. мова/Ритм.",
        "Біологія і екологія": "Біологія і еколог.",
        "Музичний інструмант": "Муз. інструмант",
        "Діяльністний підхід (ЛЕГО)": "Діяльніст. підхід (ЛЕГ.)",
        "Метод технол / обр мист з метод": "Мет. техн. / обр мист.",
        "Музичний інструмент": "Муз. інструмент",
        "Методика навч математики": "Метод. навч. мат.",
        "Інклюзивна педагогіка": "Інклюз. педаг.",
        "Інформ комун технології": "Інформ. ком. тех.",
        "Метод техн/Обр мист з метод": "Мет. техн./Обр мист.",
        "Основи укр державності": "Основи укр. держ.",
        "Метод навч англ мови": "Метод. навч. англ.",
        "Осн інклюз / Муз інструмент": "Осн інклюз / Муз інстр.",
        "Ритміка / Укр м (практикум)": "Ритміка / Укр м (практ.)",
        "Укр м (практ)/ Ритміка": "Укр м (практ.)/Ритміка",
        "Осн інклюз педагогіки": "Осн інклюз пед.",
        "Осн інкл педаго/ Пол мова": "Осн інкл педаго/Пол мова",
        "Пол мова/ Ритміка": "Пол мова/Ритміка",
        "Метод тренажу і гімнастики": "Мет. тренажу і гімн.",
        "Осн педагог майстерності": "Осн пед. майст.",
        "Англ мова / Інформатика": "Англ мова / Інформ.",
        "Осн правознавства": "Осн правознав.",
        "Істоія: Україна і світ": "Істоія: Укр і світ",
        "Метод навч математики": "Метод. навч. мат.",
        "Практ курс англ мови": "Практ. курс англ.",
        "Метод навч укр мови": "Метод. навч. укр.",
        "Обр з метод/ Метод технолог": "Обр з метод/Мет. техн.",
        "Практ курс іноз мови": "Практ. курс іноз.",
        "Прак курс іноз мови": "Прак курс іноз.",
        "Мтеор і метод музики": "Мтеор. і метод муз.",
        "Фізичне виховання": "Фізичне вих.",
        "Культура мовл і дит літ": "Культура мовл. і дит.",
        "Практ курс англ мови (І)": "Практ. курс англ (І)",
        "Польська / Ритміка": "Польська / Ритміка",
        "Ритміка / Польська": "Ритміка / Польська",
        "Вибіркові дисципліни": "Вибірк. дисц.",
        "Практикум української мови": "Практикум укр. мови",
        "Осн інклюзивної педагог": "Осн інклюз. пед.",
        "Осн науков досліджень": "Осн наук. досл.",
        "Дитяча література": "Дитяча літ.",
        "Комун інтенсив з англ мови": "Комун інт. з англ.",
        "Основи роботи з ЕОМ": "Осн. роботи з ЕОМ",
        "Інф системи і мережі": "Інф. системи і мер.",
        "Економічна теорія": "Економ. теорія",
        "Метод проф освіти": "Мет. проф. освіти",
        "Прикладне програмування": "Прикл. програмування",
        "Прикладне програмування та веб-дизайн": "Прикл. прогр. та веб-диз.",
        "Інженерна графіка": "Інженерна граф.",
        "Архітектура комп'ютерів": "Арх. комп'ютерів",
        "Мови програмування": "Мови прогр.",
        "Осн укр державності": "Осн укр. держ.",
        "Вибіркова дсиципліна": "Вибіркова дсц.",
        "Захист інформації": "Захист інф.",
        "Контроль діагностика ПК": "Контроль діагн. ПК",
        "Історія : Україна і світ": "Історія: Укр і світ",
        "Нарисна геометрія": "Нарисна геом.",
        "Основи виробництва": "Осн. виробницт.",
        "ІКТ та програмування": "ІКТ та прогр.",
        "Креслення/Осн вир": "Креслення/Осн. вир",
        "Осн вироб/Креслення": "Осн. вироб/Кресл.",
        "ІКТ та техн програмування": "ІКТ та техн. прогр.",
        "Практика удосконалення навичок": "Практика удоск. нав.",
        "Укр кухня/ Рац харч (виб)": "Укр кухня/Рац харч",
        "Інф-комунік технології": "Інф-комунік тех.",
        "Іноземна мова (проф)": "Іноз. мова (проф)",
        "Мет навч технології/Осн виробництва": "Мет. навч. техн./Осн. вироб.",
        "Осн вироб/ Мет навч": "Осн. вироб/Мет. навч.",
        "Технологіч практикум": "Техн. практикум",
        "Прикладне прогрмування": "Прикл. прогр.",
        "Прик програм/ Методика": "Прик прогр./Метод.",
        "Методика /Рац харчуван": "Метод. /Рац харч.",
        "Осн електротехніки": "Осн. електротех.",
        "Осн наукового дослідження": "Осн. наук. досл.",
        "Польська мова (вибіркова)": "Польська мова (виб.)",
        "Технологічний практикум": "Техн. практикум",
        "Прикл прогр / Будова двиг": "Прикл. прогр. / Будова двиг.",
        "Методика технологій": "Метод. техн.",
        "Вибіркова дисц ООД": "Вибіркова дисц.",
        "Економіка підприємства": "Економіка підпр.",
        "Польська / Бібл справа": "Польська / Бібл. справа",
        "Соц мережі та управл репутац/НПВ": "Соц мережі та уп. реп.",
        "Осн менеджменту і маркетингу": "Осн. менеджмент і маркет.",
        "Пол мова/Бібл справа": "Пол мова/Бібл. справа",
        "Етика діл спілкування": "Етика діл. спіл.",
        "Соц мережі та упр репутацією": "Соц мережі та уп. реп.",
        "Тер звязків з гром": "Тер звязків з гром.",
        "Укр діл мовл та редагув": "Укр діл мовл та ред",
        "Укр ділове мовл і редагування": "Укр діл мовл і ред",
        "Осн. укр. державності": "Осн. укр. держ",
        "Інф забезп управління": "Інф забезп уп",
        "Теор і прак зв'язків з громад": "Теор і прак зв'язк",
        "Осн сценічного мист": "Осн сценічн мист",
        "Інформ забез управління": "Інформ забез уп",
        "Історія (Україна і світ)": "Історія (Укр і світ)",
        "Історія: Україна світ": "Історія: Укр світ",
        "Вибіркова/   Ритміка": "Вибіркова/ Ритміка",
        "Ритміка/ ДПМ (виб)": "Ритміка/ ДПМ",
        "Англ мова (І підгр)": "Англ мова (І п)",
        "Осн природи з метод": "Осн природи з м",
        "Осн пед майстерності": "Осн пед майстер",
        "Метод образтв діяльн": "Метод образтв діят",
        "Осн інкл/Практ/Обр мист курс": "Осн інкл/Практ/Обр",
        "Метод озн з природою": "Метод озн прир",
        "Осн інкл пед/Обр мист/Іноз мов": "Осн інкл пед/Обр",
        "Практ англ / Інклюз пед": "Практ англ / Інклюз",
        "Інклюз / Практ англ": "Інклюз / Практ",
        "Людина і сусп, суспільствознавство": "Людина і сусп, сусп",
        "Дослідницька робота студентів": "Дослідницька робота",
        "Метод обр діял/Мет муз вих": "Метод обр діял/Мет муз",
        "Мет обр діял/Теор муз": "Мет обр діял/Теор муз",
        "Література для дітей": "Літер для дітей",
        "Англ / Інклюз / МІ": "Англ / Інклюз / МІ",
        "Дослідницька робота": "Дослідницька робота",
        "Метод. розвитку мовлення і логопедія": "Метод. розвитку мовл",
        "Мет фіз вих/Мет розв мов": "Мет фіз вих/Мет розв",
        "Музичний інстрімент": "Музичний інстр",
        "Обр мист / Англ мова (проф)": "Обр мист / Англ мова",
        "Англ мова / Обр мист": "Англ мова / Обр мист",
        "ІКТ/Метод. розв мовл": "ІКТ/Метод. розв",
        "Педагогіка з осн наук досл": "Педагогіка з осн наук",
        "Осн інкл/Обр мист/Практ ін м": "Осн інкл/Обр мист/Практ",
    }
    result = []
    for row in list_:
        cut_words = []
        for word in row:

            if word is not None and type(word) is not int:
                if f"{word[-1]}" == " ":
                    word = word[:-1]
            if len(f"{word}") > 16:
                try:
                    cut_words.append(change_on_it[word])
                except:
                    cut_words.append(word[:21])
            else:
                cut_words.append(word)
        result.append(cut_words)
    return result


def get_finished_schedule_object(column: int) -> Schedule:
    raw_data = get_data_from_sheet(column)
    fried_data = cut_big_words(raw_data)
    html_format = transform_list_to_html_list(fried_data)
    return Schedule.from_dict(html_format)


def get_first_block(element: str, lesson: int, schedules: Schedule):
    return f"""
                <tr>
                    <td class="number-lesson start" rowspan="2">
                        {lesson}
                    </td>
                    {getattr(schedules.monday, element).numerator}
                    {getattr(schedules.monday, element).numerator_audience}
                    {getattr(schedules.tuesday, element).numerator}
                    {getattr(schedules.tuesday, element).numerator_audience}
                    {getattr(schedules.wednesday, element).numerator}
                    {getattr(schedules.wednesday, element).numerator_audience}
                </tr>
                <tr>                    
                    {getattr(schedules.monday, element).denominator}
                    {getattr(schedules.monday, element).denominator_audience}
                    {getattr(schedules.tuesday, element).denominator}
                    {getattr(schedules.tuesday, element).denominator_audience}
                    {getattr(schedules.wednesday, element).denominator}
                    {getattr(schedules.wednesday, element).denominator_audience}
                </tr>
    """


def get_second_block(element: str, lesson: int, schedules: Schedule):
    return f"""
                <tr>
                    <td class="number-lesson start" rowspan="2">
                        {lesson}
                    </td>
                    {getattr(schedules.thursday, element).numerator}
                    {getattr(schedules.thursday, element).numerator_audience}
                    {getattr(schedules.friday, element).numerator}
                    {getattr(schedules.friday, element).numerator_audience}
                    {getattr(schedules.monday, element).numerator}
                    {getattr(schedules.monday, element).numerator_audience}
                </tr>
                <tr>                    
                    {getattr(schedules.thursday, element).denominator}
                    {getattr(schedules.thursday, element).denominator_audience}
                    {getattr(schedules.friday, element).denominator}
                    {getattr(schedules.friday, element).denominator_audience}
                    {getattr(schedules.monday, element).denominator}
                    {getattr(schedules.monday, element).denominator_audience}
                </tr>
    """


def get_theme() -> dict:
    return {
        "black": {
            "background-color": "white",
            "start": "#000000",
            "week-name": "#2e2e2e",
            "light": "#E8E8E8",
            "dark": "#CBCBCB",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "#989898FF",
        },
        "gray": {
            "background-color": "white",
            "start": "#A6A6A6",
            "week-name": "#A6A6A6",
            "light": "#E8E8E8",
            "dark": "#CBCBCB",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "#989898FF",
        },
        "red": {
            "background-color": "white",
            "start": "#5b0000",
            "week-name": "#9b0000",
            "light": "#ffbdbd",
            "dark": "#ffa1a1",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "#E15B5BFF",
        },
        "orange": {
            "background-color": "white",
            "start": "#b95102",
            "week-name": "#d65900",
            "light": "#ffb692",
            "dark": "#fb9b63",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "#ff9840",
        },
        "purple": {
            "background-color": "white",
            "start": "#45007b",
            "week-name": "#7701d1",
            "light": "#e3c8ff",
            "dark": "#d0a6ff",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "#b783f9",
        },
        "pink": {
            "background-color": "white",
            "start": "#680068",
            "week-name": "#9e019e",
            "light": "#fac8fa",
            "dark": "#fb9ffb",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "#eb85ff",
        },
        "green": {
            "background-color": "white",
            "start": "#125b00",
            "week-name": "#178101",
            "light": "#cafabe",
            "dark": "#aeffa0",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "#94f868",
        },
        "brown": {
            "background-color": "white",
            "start": "#591f01",
            "week-name": "#833400",
            "light": "rgba(96, 33, 1, 0.4)",
            "dark": "rgba(96, 34, 0, 0.61)",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "#956E5BFF",
        },
        "blue": {
            "background-color": "white",
            "start": "#01144e",
            "week-name": "#00188a",
            "light": "#a8beff",
            "dark": "#7f92fb",
            "td-color-text": "black",
            "week-color-text": "white",
            "start-color-text": "white",
            "practice_start": "№6F8CF6FF",
        },
        "catppuccino": {
            "background-color": "#1e1e2e",
            "start": "#11111b",
            "week-name": "#181825",
            "light": "#45475a",
            "dark": "#313244",
            "td-color-text": "bac2de",
            "week-color-text": "bac2de",
            "start-color-text": "bac2de",
            "practice_start": "#474763",
        },
    }


def import_data_to_html(schedules: Schedule, theme: str):
    colors = get_theme()[theme]
    text = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Title</title>
    <style>
        :root {{
            --background-color: {colors["background-color"]};
            --start: {colors["start"]};
            --week-name: {colors["week-name"]};
            --light: {colors["light"]};
            --dark: {colors["dark"]};
            --td-color-text: {colors["td-color-text"]};
            --week-color-text: {colors["week-color-text"]};
            --start-color-text: {colors["start-color-text"]};
            --practice_start: {colors["practice_start"]};
        }}
    </style>
    <link rel="stylesheet" href="../../style.css">
</head>
<body>
    <div class="center">
        <table>
            <tr class="week">
                <td class="start">{schedules.group_name}</td>
                <td class="week-name" colspan="2">{schedules.monday.day}</td>
                <td class="week-name" colspan="2">{schedules.tuesday.day}</td>
                <td class="week-name" colspan="2">{schedules.wednesday.day}</td>
            </tr>

            {get_first_block("first", 1, schedules)}
            {get_first_block("second", 2, schedules)}
            {get_first_block("third", 3, schedules)}
            {get_first_block("fourth", 4, schedules)}
            {get_first_block("fifth", 5, schedules)}
            {get_first_block("sixth", 6, schedules)}


            <tr class="week">
                <td class="start">{schedules.group_name}</td>
                <td class="week-name" colspan="2">{schedules.thursday.day}</td>
                <td class="week-name" colspan="2">{schedules.friday.day}</td>
                <td class="week-name" colspan="2">Субота</td>
            </tr>
            {get_second_block("first", 1, schedules)}
            {get_second_block("second", 2, schedules)}
            {get_second_block("third", 3, schedules)}
            {get_second_block("fourth", 4, schedules)}
            {get_second_block("fifth", 5, schedules)}
            {get_second_block("sixth", 6, schedules)}
        </table>
    </div>
</body>
</html>"""
    pathlib.Path(f"./variant/{theme}").mkdir(parents=True, exist_ok=True)
    with open(f"./variant/{theme}/{schedules.group_name}.html", "w") as file:
        file.write(text)


def parse_all_schedules(count: int, theme):
    # к-ть розкладів
    for i in range(1, count + 1):
        schedules = get_finished_schedule_object(i)
        import_data_to_html(schedules, theme)


def post_process_image(screenshot):
    image = Image.open(screenshot)

    image_data = image.load()
    width, height = image.size

    cut_color = image_data[width - 1, height - 1]
    x, y = 0, 0

    for height in range(height):
        r, g, b, a = image_data[0, height]
        if (r, g, b, a) != cut_color:
            x = height

    for width in range(width):
        r, g, b, a = image_data[width, 0]
        if (r, g, b, a) != cut_color:
            y = width

    im1 = image.crop((0.0, 0.0, y, x))
    im1.save(screenshot)


def parse_all_schedules_to_photo(driver: webdriver.Firefox, theme: str) -> None:
    driver.fullscreen_window()
    groups_list = []
    group_name = glob(f"./variant/{theme}/*.html", recursive=True)
    pathlib.Path(f"./screenshots/{theme}").mkdir(parents=True, exist_ok=True)
    for i in range(len(group_name)):
        group_names = f"{group_name[i][-7:-5]}.png"
        path_to_screenshot = f"./screenshots/{theme}/{group_names}"
        html_file = f"file:///home/salo/huta/{group_name[i]}"

        groups_list.append(group_names + " ")
        driver.get(html_file)

        driver.save_screenshot(path_to_screenshot)
        post_process_image(path_to_screenshot)

    with open("index.html", "w") as file:
        file.write("".join(groups_list))


def parsing_all_themes(theme_name: str, driver: webdriver.Firefox):
    parse_all_schedules(36, theme_name)
    parse_all_schedules_to_photo(driver, theme_name)
    driver.quit()


def main():
    for theme in get_theme():
        driver = webdriver.Firefox()
        theme_process = threading.Thread(
            target=parsing_all_themes, args=(theme, driver)
        )
        theme_process.start()


if "__main__" == __name__:
    main()
